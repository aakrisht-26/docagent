"""
DocumentEditorSkill — applies natural-language edit instructions to a document
and returns a modified version as plain text / Markdown (PDF) or Excel bytes.

The LLM is given the document text and edit instructions and asked to return
the COMPLETE modified document. For Excel files, the LLM is additionally asked
to return a structured JSON edit operation that is applied with pandas/openpyxl.
"""

from __future__ import annotations

import io
import json
import time
from typing import Any, Dict, List, Optional

from core.models import SkillInput, SkillOutput
from skills.base_skill import BaseSkill
from utils.logger import get_logger

logger = get_logger(__name__)

_MAX_TEXT_CHARS = 12_000   # Token budget for raw_text in the prompt


class DocumentEditorSkill(BaseSkill):
    """
    Applies user-described edit instructions to a document via LLM.

    SkillInput.data keys:
        edit_instructions (str)              — plain English instructions
        raw_text          (str)              — cleaned document text
        file_type         (str)              — "pdf" | "excel"
        parsed_document   (ParsedDocument)   — used for Excel sheet names (optional)

    Returns SkillOutput.data with:
        modified_text   (str)   — modified document as Markdown
        output_format   (str)   — "pdf" | "excel"
        output_bytes    (bytes) — populated for Excel output
    """

    name = "document_editor"
    description = "Applies natural-language edit instructions to a document."
    required_inputs = ["edit_instructions", "raw_text", "file_type"]

    def execute(self, inputs: SkillInput) -> SkillOutput:
        start = time.monotonic()

        instructions: str  = inputs.data["edit_instructions"]
        raw_text: str      = inputs.data["raw_text"]
        file_type: str     = inputs.data.get("file_type", "pdf")
        parsed_doc         = inputs.data.get("parsed_document")

        from utils.llm_client import LLMClient
        llm = LLMClient.from_config(self.config)
        if not llm.available:
            return SkillOutput(
                success=False, data=None,
                error="Document editing requires an LLM. Set GROQ_API_KEY.",
                duration_ms=(time.monotonic() - start) * 1000,
            )

        warnings: List[str] = []
        truncated = len(raw_text) > _MAX_TEXT_CHARS
        text_for_prompt = raw_text[:_MAX_TEXT_CHARS]
        if truncated:
            warnings.append(
                f"Document truncated to {_MAX_TEXT_CHARS:,} characters for editing. "
                "Changes to content beyond this point may not be applied."
            )

        if file_type == "excel":
            return self._edit_excel(
                instructions, text_for_prompt, parsed_doc, llm, warnings, start
            )
        else:
            return self._edit_pdf(instructions, text_for_prompt, llm, warnings, start)

    # ── PDF / generic text edit ────────────────────────────────────────────────

    def _edit_pdf(
        self,
        instructions: str,
        text: str,
        llm: Any,
        warnings: List[str],
        start: float,
    ) -> SkillOutput:
        system_prompt = (
            "You are a precise document editor. "
            "You will be given a document and edit instructions. "
            "Return the COMPLETE modified document in plain text / Markdown. "
            "Apply ONLY the described changes. Do NOT omit unchanged sections. "
            "Preserve all original headings, structure, and formatting."
        )
        user_prompt = (
            f"ORIGINAL DOCUMENT:\n{text}\n\n"
            f"EDIT INSTRUCTIONS:\n{instructions}\n\n"
            "Return the full modified document below:"
        )

        modified = llm.chat(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_prompt},
            ],
            max_tokens=4000,
            temperature=0.1,
        )

        if not modified:
            return SkillOutput(
                success=False, data=None,
                error="LLM returned no response for edit.",
                duration_ms=(time.monotonic() - start) * 1000,
            )

        return SkillOutput(
            success=True,
            data={"modified_text": modified, "output_format": "pdf"},
            warnings=warnings,
            duration_ms=(time.monotonic() - start) * 1000,
        )

    # ── Excel / CSV edit ───────────────────────────────────────────────────────

    def _edit_excel(
        self,
        instructions: str,
        text: str,
        parsed_doc: Any,
        llm: Any,
        warnings: List[str],
        start: float,
    ) -> SkillOutput:
        sheet_names: List[str] = []
        if parsed_doc and hasattr(parsed_doc, "sheet_names"):
            sheet_names = parsed_doc.sheet_names or []

        sheet_hint = ""
        if sheet_names:
            sheet_hint = f"\nAvailable sheets: {', '.join(sheet_names)}. Specify which sheet each change targets."

        system_prompt = (
            "You are a precise spreadsheet editor. "
            "You will be given a text representation of a spreadsheet and edit instructions. "
            "First, return a JSON array of edit operations in this exact format:\n"
            '{"operations": [{"op": "rename_column", "sheet": "Sheet1", "old": "Name", "new": "Full Name"}, ...]}\n\n'
            "Supported ops: rename_column, set_cell_value, delete_row, add_column, rename_sheet.\n"
            "After the JSON block, also return the COMPLETE modified document in plain text / Markdown "
            "so it can be previewed even if the structured ops cannot be applied."
        )
        user_prompt = (
            f"DOCUMENT (text representation):\n{text}\n"
            f"{sheet_hint}\n\n"
            f"EDIT INSTRUCTIONS:\n{instructions}\n\n"
            "JSON operations + modified document:"
        )

        response = llm.chat(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_prompt},
            ],
            max_tokens=3000,
            temperature=0.1,
        )

        if not response:
            return SkillOutput(
                success=False, data=None,
                error="LLM returned no response for Excel edit.",
                duration_ms=(time.monotonic() - start) * 1000,
            )

        # Try to parse the JSON operations block
        output_bytes: Optional[bytes] = None
        try:
            json_match = __import__("re").search(r'\{.*?"operations".*?\}', response, __import__("re").DOTALL)
            if json_match and parsed_doc:
                ops_data = json.loads(json_match.group(0))
                output_bytes = self._apply_excel_ops(ops_data.get("operations", []), parsed_doc)
        except Exception as exc:
            logger.warning(f"Could not apply structured Excel ops: {exc}")
            warnings.append("Structured Excel edit could not be applied; preview shows text representation only.")

        # Extract the markdown part (after the JSON block)
        modified_text = response
        json_end = response.find("}")
        if json_end != -1:
            remainder = response[json_end + 1:].strip()
            if remainder:
                modified_text = remainder

        data: Dict[str, Any] = {
            "modified_text": modified_text,
            "output_format": "excel" if output_bytes else "pdf",
        }
        if output_bytes:
            data["output_bytes"] = output_bytes

        return SkillOutput(
            success=True,
            data=data,
            warnings=warnings,
            duration_ms=(time.monotonic() - start) * 1000,
        )

    @staticmethod
    def _apply_excel_ops(operations: List[Dict], parsed_doc: Any) -> Optional[bytes]:
        """Apply structured edit operations to the parsed document's data using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return None

        if not parsed_doc or not hasattr(parsed_doc, "tables"):
            return None

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        # Reconstruct sheets from tables in parsed_doc
        for table in parsed_doc.tables:
            sheet_name = str(table.get("sheet", "Sheet1"))[:31]  # Excel 31-char limit
            ws = wb.create_sheet(title=sheet_name)
            preview = table.get("preview", [])
            for row in preview:
                ws.append(row if isinstance(row, list) else [str(row)])

        # Apply each operation
        for op in operations:
            op_type   = op.get("op", "")
            sheet_ref = op.get("sheet")
            try:
                ws = wb[sheet_ref] if sheet_ref and sheet_ref in wb.sheetnames else (wb.active or next(iter(wb.worksheets), None))
                if ws is None:
                    continue

                if op_type == "rename_column":
                    old_col, new_col = op.get("old"), op.get("new")
                    for cell in ws[1]:  # header row
                        if cell.value == old_col:
                            cell.value = new_col
                            break

                elif op_type == "set_cell_value":
                    row, col, val = op.get("row"), op.get("col"), op.get("value")
                    if row and col:
                        ws.cell(row=int(row), column=int(col), value=val)

                elif op_type == "rename_sheet" and sheet_ref:
                    ws.title = str(op.get("new", sheet_ref))[:31]

                elif op_type == "add_column":
                    header = op.get("header", "New Column")
                    max_col = ws.max_column + 1
                    ws.cell(row=1, column=max_col, value=header)

            except Exception as exc:
                logger.warning(f"Could not apply op {op_type}: {exc}")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
