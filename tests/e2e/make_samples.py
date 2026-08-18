"""Regenerate the e2e sample files under tests/e2e/samples/.

The samples are committed to the repo, so a fresh clone does NOT need to run
this. It exists so the fixtures can be rebuilt or adjusted deterministically.

Usage:
    python tests/e2e/make_samples.py

Notes:
    - sample_report.pdf and sample_sales.xlsx are built from reportlab/openpyxl
      and are fully cross-platform.
    - sample_audio.wav is synthesised with the Windows SAPI speech engine, so it
      can only be regenerated on Windows. The committed copy is used everywhere
      else; this script leaves it untouched if it cannot be rebuilt.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

SAMPLES = Path(__file__).resolve().parent / "samples"
SAMPLES.mkdir(parents=True, exist_ok=True)


# ── PDF ───────────────────────────────────────────────────────────────────────

def build_pdf() -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    out = SAMPLES / "sample_report.pdf"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out), pagesize=LETTER)
    story = [
        Paragraph("Acme Robotics — Q3 Engineering Report", styles["Title"]),
        Spacer(1, 12),
        Paragraph("1. Executive Summary", styles["Heading1"]),
        Paragraph(
            "Acme Robotics completed the third quarter with the Orion control platform reaching "
            "general availability. Total engineering headcount grew from 42 to 51. The platform now "
            "serves 1,280 industrial units across 14 customer sites. Mean time between failures "
            "improved from 410 hours to 690 hours following the firmware 4.2 rollout. "
            "Cloud infrastructure spend was 312,000 dollars, under the 350,000 dollar budget.",
            styles["BodyText"]),
        Spacer(1, 12),
        Paragraph("2. Reliability Metrics", styles["Heading1"]),
        Paragraph(
            "The firmware 4.2 release addressed the servo calibration drift reported by three "
            "customers in Q2. Post-deployment telemetry shows drift incidents fell by 87 percent. "
            "The remaining incidents cluster in units manufactured before serial number A-8800.",
            styles["BodyText"]),
        Spacer(1, 12),
    ]

    table = Table([
        ["Metric", "Q2", "Q3", "Change"],
        ["MTBF (hours)", "410", "690", "+68%"],
        ["Active units", "940", "1280", "+36%"],
        ["Drift incidents", "31", "4", "-87%"],
        ["Cloud spend (USD)", "298,000", "312,000", "+4.7%"],
        ["Headcount", "42", "51", "+21%"],
    ], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    story.append(table)
    story.append(PageBreak())
    story.append(Paragraph("3. Risks and Next Quarter", styles["Heading1"]))
    story.append(Paragraph(
        "The principal risk for Q4 is the single-source supplier for the LIDAR module. "
        "A qualification programme for a second supplier began in September and is expected "
        "to complete in January. Until then a 6-week buffer stock is held. "
        "Engineering will prioritise the multi-tenant scheduler and the on-premise deployment "
        "option requested by two enterprise customers.",
        styles["BodyText"]))
    doc.build(story)
    return out


# ── Excel ─────────────────────────────────────────────────────────────────────

def build_excel() -> Path:
    from openpyxl import Workbook

    out = SAMPLES / "sample_sales.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3 Sales"
    ws.append(["Region", "Product", "Units", "Unit Price", "Revenue"])
    for row in [
        ("North", "Orion Controller", 320, 1250, 400000),
        ("North", "LIDAR Module", 180, 890, 160200),
        ("South", "Orion Controller", 210, 1250, 262500),
        ("South", "Service Contract", 95, 4200, 399000),
        ("EMEA", "Orion Controller", 410, 1180, 483800),
        ("EMEA", "LIDAR Module", 260, 850, 221000),
        ("APAC", "Orion Controller", 140, 1310, 183400),
        ("APAC", "Service Contract", 60, 4500, 270000),
    ]:
        ws.append(list(row))

    ws2 = wb.create_sheet("Headcount")
    ws2.append(["Department", "Q2", "Q3", "Open Roles"])
    for row in [("Engineering", 42, 51, 6), ("Sales", 18, 21, 3),
                ("Support", 12, 15, 2), ("Operations", 9, 9, 0)]:
        ws2.append(list(row))

    wb.save(out)
    return out


# ── Audio (Windows only) ──────────────────────────────────────────────────────

_SPEECH = (
    "Quarterly operations briefing. Revenue for the third quarter reached four point two "
    "million dollars, an increase of eighteen percent over the prior quarter. The engineering "
    "team shipped the new authentication service and reduced average API latency from three "
    "hundred milliseconds to ninety milliseconds. Customer churn fell to two point one percent. "
    "The primary risk for the next quarter is supply chain delay affecting hardware deliveries "
    "in November."
)


def build_audio() -> Path | None:
    out = SAMPLES / "sample_audio.wav"
    if sys.platform != "win32":
        print(f"  skip  {out.name} (Windows SAPI only; committed copy retained)")
        return None

    script = (
        "Add-Type -AssemblyName System.Speech; "
        "$s = New-Object System.Speech.Synthesis.SpeechSynthesizer; "
        f"$s.SetOutputToWaveFile('{out}'); "
        "$s.Rate = -1; "
        f"$s.Speak('{_SPEECH}'); "
        "$s.Dispose()"
    )
    subprocess.run(["powershell", "-NoProfile", "-Command", script], check=True)
    return out


# ── Scanned PDF (image-only, no text layer) ───────────────────────────────────

def build_scanned_pdf() -> Path:
    """Rasterise sample_report.pdf into an image-only PDF with no text layer.

    This is the fixture that forces the OCR tier: pdfplumber and PyMuPDF both
    return empty text for it, so PDFReaderSkill escalates to Tesseract.

    Mild, deterministic scan artefacts are applied so the OCR preprocessing is
    actually exercised rather than handed a pristine render:
      - 150 DPI, the typical output of a consumer scanner
      - a 1.2 degree skew, inside the 0.5-15 degree window the deskew step handles
      - a brightness gradient, which is what adaptive thresholding exists for
      - light Gaussian noise
    """
    import cv2
    import fitz
    import numpy as np

    source = SAMPLES / "sample_report.pdf"
    if not source.exists():
        build_pdf()

    out = SAMPLES / "sample_scanned.pdf"
    rng = np.random.default_rng(20260729)  # fixed seed keeps the fixture stable

    src = fitz.open(str(source))
    dst = fitz.open()

    for page in src:
        pix = page.get_pixmap(dpi=150)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY) if pix.n == 3 else img[:, :, 0]

        # Skew, filling with white so the page edges stay paper-coloured
        h, w = gray.shape[:2]
        matrix = cv2.getRotationMatrix2D((w // 2, h // 2), 1.2, 1.0)
        gray = cv2.warpAffine(gray, matrix, (w, h), flags=cv2.INTER_CUBIC,
                              borderMode=cv2.BORDER_CONSTANT, borderValue=255)

        # Uneven illumination: darker down the left edge, like a book scan
        gradient = np.linspace(0.88, 1.0, w, dtype=np.float32)[None, :]
        gray = np.clip(gray.astype(np.float32) * gradient, 0, 255)

        # Light sensor noise. Kept low on purpose: this fixture is a regression
        # baseline, so it should look like a real scanner rather than an
        # adversarial one. Heavier noise made Tesseract misread "14 customer
        # sites" as "414", which then propagated into the summary and made the
        # stage useless for spotting genuine regressions.
        gray = np.clip(gray + rng.normal(0, 2.0, gray.shape), 0, 255).astype(np.uint8)

        # JPEG, as a real scanner would produce, and ~13x smaller than PNG here
        ok, buf = cv2.imencode(".jpg", gray, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
        if not ok:
            raise RuntimeError("failed to encode scanned page")

        new_page = dst.new_page(width=page.rect.width, height=page.rect.height)
        new_page.insert_image(new_page.rect, stream=buf.tobytes())

    dst.save(str(out))
    dst.close()
    src.close()
    return out


# ── Large fixtures for RAG retrieval evaluation ───────────────────────────────

def build_large_pdf() -> Path:
    """A 20-page PDF where retrieval is a genuine decision.

    Page-level chunking yields 20 candidates against a top-3 selection, so
    unlike the small fixtures the selector must actually choose. The page
    content is designed to be hard — see tests/e2e/rag_eval/fixture_content.py
    for the three difficulties it builds in.
    """
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer)

    sys.path.insert(0, str(Path(__file__).resolve().parent / "rag_eval"))
    from fixture_content import LARGE_PDF_PAGES, LARGE_PDF_TITLE

    out = SAMPLES / "sample_large_report.pdf"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out), pagesize=LETTER)

    story = []
    for index, (heading, body) in enumerate(LARGE_PDF_PAGES):
        if index == 0:
            story.append(Paragraph(LARGE_PDF_TITLE, styles["Title"]))
            story.append(Spacer(1, 12))
        story.append(Paragraph(heading, styles["Heading1"]))
        story.append(Paragraph(body, styles["BodyText"]))
        if index != len(LARGE_PDF_PAGES) - 1:
            story.append(PageBreak())   # one section per page, so page == chunk

    doc.build(story)
    return out


def build_large_excel() -> Path:
    """An 8-sheet workbook where sheet-level chunking makes retrieval matter."""
    from openpyxl import Workbook

    sys.path.insert(0, str(Path(__file__).resolve().parent / "rag_eval"))
    from fixture_content import LARGE_XLSX_SHEETS

    out = SAMPLES / "sample_large_sales.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, (header, rows) in LARGE_XLSX_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(list(header))
        for row in rows:
            ws.append(list(row))
    wb.save(out)
    return out


def build_dense_pdf() -> Path:
    """An 8-page PDF whose pages each carry six unrelated topics.

    The other large fixture averages 46 words per page with one topic each, so
    a page is already passage-sized there and sub-chunking cannot change the
    outcome either way. These pages average ~214 words across six unrelated
    subjects, which is the condition under which retrieving a whole page
    plausibly dilutes the match. See fixture_content.py for the design.
    """
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer)

    sys.path.insert(0, str(Path(__file__).resolve().parent / "rag_eval"))
    from fixture_content import DENSE_PDF_PAGES, DENSE_PDF_TITLE

    out = SAMPLES / "sample_dense_manual.pdf"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out), pagesize=LETTER)

    story = []
    for index, (heading, body) in enumerate(DENSE_PDF_PAGES):
        if index == 0:
            story.append(Paragraph(DENSE_PDF_TITLE, styles["Title"]))
            story.append(Spacer(1, 12))
        story.append(Paragraph(heading, styles["Heading1"]))
        story.append(Paragraph(body, styles["BodyText"]))
        if index != len(DENSE_PDF_PAGES) - 1:
            story.append(PageBreak())   # one section per page, so page == chunk

    doc.build(story)
    return out


def build_questionnaire_pdf() -> Path:
    """A form the HEURISTIC does not recognise, so only the LLM blend can.

    Every other fixture is a normal document, so question_extraction and the
    `0.7 * llm_score + 0.3 * heuristic` blend had never run end to end. See the
    note above QUESTIONNAIRE_PDF_PAGES for why it is written in prose rather
    than with the usual form furniture.
    """
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer)

    sys.path.insert(0, str(Path(__file__).resolve().parent / "rag_eval"))
    from fixture_content import QUESTIONNAIRE_PDF_PAGES, QUESTIONNAIRE_PDF_TITLE

    out = SAMPLES / "sample_questionnaire.pdf"
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(out), pagesize=LETTER)

    story = []
    for index, (heading, body) in enumerate(QUESTIONNAIRE_PDF_PAGES):
        if index == 0:
            story.append(Paragraph(QUESTIONNAIRE_PDF_TITLE, styles["Title"]))
            story.append(Spacer(1, 12))
        story.append(Paragraph(heading, styles["Heading1"]))
        for line in body.splitlines():
            if line.strip():
                story.append(Paragraph(line, styles["BodyText"]))
        if index != len(QUESTIONNAIRE_PDF_PAGES) - 1:
            story.append(PageBreak())

    doc.build(story)
    return out


if __name__ == "__main__":
    for path in (build_pdf(), build_excel(), build_audio(), build_scanned_pdf(),
                 build_large_pdf(), build_large_excel(), build_dense_pdf(),
                 build_questionnaire_pdf()):
        if path is not None:
            print(f"  wrote {path.name:24s} {path.stat().st_size / 1024:8.1f} KB")
    print(f"\nSamples in {SAMPLES}")
